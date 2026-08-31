"""
test_phase2.py — Script de prueba para la Fase 2: Modulo de Escritura.

Flujo completo:
  1. [Fase 1] Extrae datos de ENERO de cada archivo en origen_unidades/
     usando el mapper y extractor de src.core.
  2. [Fase 2] Escribe los datos extraidos en las pestanas correspondientes
     de DATATEST/destino/POA_Destino_Test.xlsx via excel_writer.
  3. [Validacion] Verifica celda a celda que los valores escritos coincidan
     exactamente con los extraidos, sin alterar encabezados ni PROG.

Uso:
    python src/writers/test_phase2.py
    # o como modulo:
    python -m src.writers.test_phase2
"""

from __future__ import annotations

import logging
import os
import sys
from pathlib import Path

# Permitir ejecucion directa sin instalar el paquete
ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

from src.core.matcher import build_mapping
from src.core.extractor import extract_month_data
from src.writers.excel_writer import write_month_data_to_excel, verify_written_data

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------
DATATEST_ROOT = ROOT / "ARCHIVOSPRUEBAS" / "DATATEST"
ORIGEN_DIR    = DATATEST_ROOT / "origen_unidades"
DESTINO_FILE  = DATATEST_ROOT / "destino" / "POA_Destino_Test.xlsx"
TARGET_MONTH  = "ENERO"

logging.basicConfig(
    level=logging.INFO,
    format="  [%(levelname)s] %(message)s",
)

SEP  = "=" * 72
SEP2 = "-" * 72
TICK = "OK"
CROSS = "FALLO"


def snapshot_real_column(dest_path: Path, sheet_name: str, month: str) -> dict:
    """
    Lee el estado actual de la columna REAL para un mes/pestana.
    Devuelve {indicador: valor}.
    """
    from src.writers.excel_writer import _find_real_column
    wb = openpyxl.load_workbook(dest_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    real_col = _find_real_column(ws, month)
    snap = {}
    for row in range(6, ws.max_row + 1):
        ind = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=real_col).value
        if ind is not None:
            try:
                snap[int(ind)] = val
            except (TypeError, ValueError):
                pass
    wb.close()
    return snap


def snapshot_prog_column(dest_path: Path, sheet_name: str, month: str) -> dict:
    """
    Lee el estado actual de la columna PROG para verificar que no cambia.
    """
    from src.writers.excel_writer import _find_real_column
    import openpyxl

    wb = openpyxl.load_workbook(dest_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    # PROG es la columna inmediatamente anterior a REAL del mismo bloque de mes
    real_col = _find_real_column(ws, month)
    prog_col = real_col - 1  # PROG siempre es la anterior a REAL
    snap = {}
    for row in range(6, ws.max_row + 1):
        ind = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=prog_col).value
        if ind is not None:
            try:
                snap[int(ind)] = val
            except (TypeError, ValueError):
                pass
    wb.close()
    return snap


def run():
    print()
    print(SEP)
    print("  DATA-TRAVEL — Fase 2: Modulo de Escritura — Prueba de integracion")
    print(SEP)
    print(f"  Origen  : {ORIGEN_DIR}")
    print(f"  Destino : {DESTINO_FILE}")
    print(f"  Mes     : {TARGET_MONTH}")
    print()

    # -----------------------------------------------------------------------
    # 1. LEER PESTANAS DESTINO
    # -----------------------------------------------------------------------
    wb_dest = openpyxl.load_workbook(DESTINO_FILE, read_only=True)
    sheet_names = wb_dest.sheetnames
    wb_dest.close()
    print(f"  Pestanas en destino ({len(sheet_names)}): {sheet_names}")
    print()

    # -----------------------------------------------------------------------
    # 2. LEER Y EMPAREJAR ARCHIVOS ORIGEN
    # -----------------------------------------------------------------------
    source_files = sorted(
        f.name for f in ORIGEN_DIR.iterdir()
        if f.suffix.lower() in {".xlsx", ".xls"}
    )
    mapping = build_mapping(source_files, sheet_names, score_cutoff=60.0)

    print(SEP)
    print("  PASO 1 — Mapeo Origen -> Destino (heredado de Fase 1)")
    print(SEP)
    for m in mapping:
        st = TICK if m["matched"] else CROSS
        print(f"  [{st}] {m['file']:<45} -> {m['sheet'] or '(sin mapeo)':<14} score={m['score']:.1f}")
    print()

    # -----------------------------------------------------------------------
    # 3. EXTRAER DATOS DE ENERO
    # -----------------------------------------------------------------------
    extractions: list[dict] = []
    for m in mapping:
        if not m["matched"]:
            continue
        filepath = ORIGEN_DIR / m["file"]
        data = extract_month_data(filepath, TARGET_MONTH)
        extractions.append({**m, "data": data})

    print(SEP)
    print(f"  PASO 2 — Extraccion de datos ({TARGET_MONTH})")
    print(SEP)
    for e in extractions:
        print(f"  Archivo : {e['file']}")
        print(f"  Pestana : {e['sheet']}  |  Indicadores extraidos: {len(e['data'])}")
        print(f"  Datos   : {e['data']}")
        print(SEP2)
    print()

    # -----------------------------------------------------------------------
    # 4. SNAPSHOT ANTES DE ESCRIBIR (para comparar)
    # -----------------------------------------------------------------------
    snapshots_before: dict[str, dict] = {}
    prog_snapshots:   dict[str, dict] = {}
    for e in extractions:
        sname = e["sheet"]
        snapshots_before[sname] = snapshot_real_column(DESTINO_FILE, sname, TARGET_MONTH)
        prog_snapshots[sname]   = snapshot_prog_column(DESTINO_FILE, sname, TARGET_MONTH)

    print(SEP)
    print("  PASO 3 — Estado ANTES de escritura (col REAL del destino)")
    print(SEP)
    for sname, snap in snapshots_before.items():
        empty_count = sum(1 for v in snap.values() if v is None)
        print(f"  Pestana '{sname}': {len(snap)} indicadores, {empty_count} celdas REAL vacias")
    print()

    # -----------------------------------------------------------------------
    # 5. ESCRIBIR EN EXCEL LOCAL
    # -----------------------------------------------------------------------
    print(SEP)
    print("  PASO 4 — Escritura en Excel local")
    print(SEP)

    results_write: list[dict] = []
    for e in extractions:
        sname = e["sheet"]
        print(f"  Escribiendo pestana '{sname}' ({len(e['data'])} valores)...", end=" ")
        ok = write_month_data_to_excel(
            dest_file_path=DESTINO_FILE,
            sheet_name=sname,
            month=TARGET_MONTH,
            data_dict=e["data"],
            create_backup=True,
        )
        status = TICK if ok else CROSS
        print(f"[{status}]")
        results_write.append({"sheet": sname, "ok": ok, "data": e["data"]})
    print()

    # -----------------------------------------------------------------------
    # 6. VALIDACION CELDA A CELDA
    # -----------------------------------------------------------------------
    print(SEP)
    print("  PASO 5 — Validacion post-escritura (celda a celda)")
    print(SEP)

    all_pass = True
    for r in results_write:
        sname = r["sheet"]
        expected = r["data"]
        verification = verify_written_data(
            DESTINO_FILE, sname, TARGET_MONTH, expected
        )
        ok_flag = verification["ok"]
        if not ok_flag:
            all_pass = False

        status = TICK if ok_flag else CROSS
        print(f"\n  [{status}] Pestana: '{sname}'")
        print(f"       Coincidencias : {len(verification['matches'])}/{verification['total_checked']}")

        if verification["mismatches"]:
            print(f"       DISCREPANCIAS  : {verification['mismatches']}")
        if verification["missing_indicators"]:
            print(f"       IND NO EN HOJA : {verification['missing_indicators']}")

        # Mostrar tabla de resultados
        print(f"       {'IND':>4}  {'ESPERADO':>10}  {'LEIDO':>10}  {'MATCH?':>7}")
        print(f"       {'-'*4}  {'-'*10}  {'-'*10}  {'-'*7}")
        for ind, exp_val in expected.items():
            if ind in verification["matches"]:
                got_val = exp_val
                match_str = "OK"
            elif ind in verification["mismatches"]:
                got_val = verification["mismatches"][ind]["got"]
                match_str = "FALLO"
                all_pass = False
            else:
                got_val = "(N/A)"
                match_str = "MISS"
            print(f"       {ind:>4}  {str(exp_val):>10}  {str(got_val):>10}  {match_str:>7}")

    # Verificar que PROG no fue alterado
    print()
    print(SEP)
    print("  PASO 6 — Verificacion de integridad (PROG no debe cambiar)")
    print(SEP)
    prog_integrity_ok = True
    for r in results_write:
        sname = r["sheet"]
        prog_after = snapshot_prog_column(DESTINO_FILE, sname, TARGET_MONTH)
        prog_before = prog_snapshots[sname]
        changed = {ind: (prog_before[ind], prog_after[ind])
                   for ind in prog_before
                   if prog_before[ind] != prog_after.get(ind)}
        if changed:
            prog_integrity_ok = False
            all_pass = False
            print(f"  [FALLO] Pestana '{sname}': PROG modificado en indicadores: {changed}")
        else:
            print(f"  [OK]    Pestana '{sname}': PROG sin cambios ({len(prog_before)} filas verificadas)")

    # -----------------------------------------------------------------------
    # 7. RESUMEN FINAL
    # -----------------------------------------------------------------------
    print()
    print(SEP)
    print("  RESUMEN FINAL DE LA FASE 2")
    print(SEP)
    pestanas_ok = sum(1 for r in results_write if r["ok"])
    print(f"  Pestanas escritas con exito : {pestanas_ok}/{len(results_write)}")
    print(f"  Validacion celda a celda    : {'PASS' if all_pass else 'FAIL'}")
    print(f"  Integridad PROG             : {'OK' if prog_integrity_ok else 'ALTERADO'}")
    print(f"  Backup generado             : {DESTINO_FILE.with_suffix('.backup.xlsx').name}")

    if all_pass and prog_integrity_ok:
        print()
        print("  [PASS] Fase 2 completada al 100%. Escritura Excel verificada.")
    else:
        print()
        print("  [WARN] Hay discrepancias. Revisar los detalles anteriores.")

    print(SEP)
    print()


if __name__ == "__main__":
    run()
