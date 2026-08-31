"""
excel_writer.py — Fase 2: Escritura de datos mensuales en un libro Excel local.

Responsabilidad:
- Recibir un diccionario {numero_indicador: valor_real} y volcarlo en la columna
  REAL del mes indicado, dentro de la pestana correcta del libro destino.
- Proteger el archivo original con backup antes de cualquier modificacion.
- Manejar graciosamente el caso en que el archivo este abierto en Excel
  (PermissionError) con un mensaje claro para el usuario final.

Estructura asumida del libro destino (identica al origen):
  Fila 4 : encabezados de mes   -> col C/E/...: "ENERO", "FEBRERO", ...
  Fila 5 : sub-encabezados      -> "PROG", "REAL", "%"
  Fila 6+: datos                -> col A: numero_indicador (int)
"""

from __future__ import annotations

import logging
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Union

import openpyxl

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers internos (reutiliza logica de extractor sin importarlo directamente)
# ---------------------------------------------------------------------------

def _strip_accents(text: str) -> str:
    """Quita diacriticos de un string."""
    nfd = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in nfd if unicodedata.category(ch) != "Mn")


def _find_real_column(ws, target_month: str) -> int:
    """
    Localiza el indice de columna (1-based) de la sub-columna REAL para el
    mes indicado.

    Algoritmo:
      1. Recorre la fila 4 buscando la celda cuyo valor (normalizado) coincida
         con target_month.
      2. Desde esa columna en adelante, busca "REAL" en la fila 5 dentro del
         bloque del mes (hasta que aparezca otra celda no-None en fila 4).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Hoja de calculo del libro destino (modo lectura/escritura).
    target_month : str
        Nombre del mes, ej: "ENERO", "FEBRERO".

    Returns
    -------
    int
        Indice de columna 1-based de la celda REAL.

    Raises
    ------
    ValueError
        Si el mes o la sub-columna REAL no se encuentran.
    """
    target = _strip_accents(target_month.strip().upper())
    max_col = ws.max_column

    # --- Paso 1: fila 4 -> columna de inicio del mes ---
    month_col_start: int | None = None
    for col in range(1, max_col + 1):
        val = ws.cell(row=4, column=col).value
        if val and _strip_accents(str(val).strip().upper()) == target:
            month_col_start = col
            break

    if month_col_start is None:
        raise ValueError(
            f"Mes '{target_month}' no encontrado en la fila 4. "
            f"Verifica que el encabezado exista en la hoja."
        )

    # --- Paso 2: fila 5 -> sub-columna REAL ---
    real_col: int | None = None
    for col in range(month_col_start, max_col + 2):
        # Salir si entramos en otro mes
        if col > month_col_start:
            next_h = ws.cell(row=4, column=col).value
            if next_h is not None:
                break
        sub = ws.cell(row=5, column=col).value
        if sub and str(sub).strip().upper() == "REAL":
            real_col = col
            break

    if real_col is None:
        raise ValueError(
            f"Sub-columna 'REAL' no encontrada para el mes '{target_month}'. "
            f"Verifica que la fila 5 contenga 'REAL' en el bloque del mes."
        )

    return real_col


def _make_backup(filepath: Path) -> Path:
    """
    Crea una copia de seguridad del archivo antes de modificarlo.

    El backup se guarda junto al original con sufijo ``.backup.xlsx``.
    Si ya existe un backup previo, se sobreescribe.

    Parameters
    ----------
    filepath : Path
        Ruta al archivo original.

    Returns
    -------
    Path
        Ruta al archivo de backup creado.
    """
    backup_path = filepath.with_suffix(".backup.xlsx")
    shutil.copy2(filepath, backup_path)
    log.info("Backup creado: %s", backup_path)
    return backup_path


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------

def write_month_data_to_excel(
    dest_file_path: Union[str, Path],
    sheet_name: str,
    month: str,
    data_dict: dict,
    create_backup: bool = True,
    header_row: int = 4,
    data_start_row: int = 6,
    indicator_col: int = 1,
) -> bool:
    """
    Escribe los valores REAL del diccionario en la hoja Excel destino.

    El libro se abre en modo lectura/escritura con openpyxl (sin alterar
    formatos, formulas ni otros datos). Solo se modifican las celdas de la
    columna REAL que correspondan a indicadores presentes en data_dict.

    Parameters
    ----------
    dest_file_path : str | Path
        Ruta al libro Excel destino (.xlsx).
    sheet_name : str
        Nombre de la pestana donde se escribiran los datos.
    month : str
        Nombre del mes objetivo, ej: ``"ENERO"``, ``"FEBRERO"``.
    data_dict : dict
        Diccionario ``{int_indicador: valor_real}`` proveniente de la
        fase de extraccion.
    create_backup : bool
        Si True (defecto), crea un backup ``.backup.xlsx`` antes de guardar.
    header_row : int
        Fila (1-based) de encabezados de mes. Default: 4.
    data_start_row : int
        Primera fila de datos (1-based). Default: 6.
    indicator_col : int
        Columna (1-based) que contiene el numero de indicador. Default: 1.

    Returns
    -------
    bool
        True si la escritura fue exitosa, False si hubo algun error.

    Raises
    ------
    FileNotFoundError
        Si el archivo destino no existe.
    KeyError
        Si la pestana indicada no existe en el libro.
    ValueError
        Si el mes o la columna REAL no se encuentran.

    Examples
    --------
    >>> ok = write_month_data_to_excel(
    ...     "POA_Destino_Test.xlsx",
    ...     sheet_name="ElCuco",
    ...     month="ENERO",
    ...     data_dict={1: 4, 2: 12, 3: 80},
    ... )
    >>> assert ok is True
    """
    dest_path = Path(dest_file_path)

    # --- Validaciones previas ---
    if not dest_path.exists():
        raise FileNotFoundError(f"Archivo destino no encontrado: {dest_path}")

    # --- Backup preventivo ---
    backup_path: Path | None = None
    if create_backup:
        backup_path = _make_backup(dest_path)

    try:
        wb = openpyxl.load_workbook(dest_path)
    except PermissionError:
        msg = (
            f"\n[ERROR] No se puede abrir '{dest_path.name}'.\n"
            "El archivo parece estar abierto en Excel u otro programa.\n"
            "Por favor cierra el archivo y vuelve a ejecutar el proceso.\n"
        )
        log.error(msg)
        print(msg)
        return False
    except Exception as exc:
        log.error("Error abriendo el libro: %s", exc)
        raise

    # --- Validar que la pestana exista ---
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise KeyError(
            f"La pestana '{sheet_name}' no existe en el libro.\n"
            f"Pestanas disponibles: {available}"
        )

    ws = wb[sheet_name]

    # --- Localizar columna REAL ---
    real_col = _find_real_column(ws, month)
    log.debug("Columna REAL para '%s' en '%s': col %d", month, sheet_name, real_col)

    # --- Escribir valores ---
    written = 0
    skipped = 0
    not_found = list(data_dict.keys())  # indicadores aun no encontrados

    for row in range(data_start_row, ws.max_row + 1):
        ind_val = ws.cell(row=row, column=indicator_col).value
        if ind_val is None:
            continue
        try:
            ind_key = int(ind_val)
        except (TypeError, ValueError):
            continue

        if ind_key in data_dict:
            ws.cell(row=row, column=real_col).value = data_dict[ind_key]
            written += 1
            if ind_key in not_found:
                not_found.remove(ind_key)
        else:
            skipped += 1

    log.info(
        "Pestana '%s' | Mes '%s' | Escritos: %d | Sin dato (skipped): %d | "
        "Indicadores del dict no encontrados en hoja: %s",
        sheet_name, month, written, skipped, not_found,
    )

    # --- Guardar el libro ---
    try:
        wb.save(dest_path)
    except PermissionError:
        msg = (
            f"\n[ERROR] No se pudo guardar '{dest_path.name}'.\n"
            "El archivo fue abierto en Excel mientras se procesaba.\n"
            "Cierra el archivo y vuelve a intentarlo.\n"
        )
        log.error(msg)
        print(msg)
        if backup_path and backup_path.exists():
            print(f"Tu backup sin modificar esta en: {backup_path}")
        return False

    wb.close()

    summary = {
        "sheet": sheet_name,
        "month": month,
        "real_col": real_col,
        "written": written,
        "skipped_rows": skipped,
        "indicators_not_in_sheet": not_found,
        "backup": str(backup_path) if backup_path else None,
    }
    log.info("Escritura completada: %s", summary)
    return True


def verify_written_data(
    dest_file_path: Union[str, Path],
    sheet_name: str,
    month: str,
    expected_dict: dict,
    data_start_row: int = 6,
    indicator_col: int = 1,
) -> dict:
    """
    Lee el archivo ya modificado y verifica que los valores escritos
    coincidan con expected_dict.

    Parameters
    ----------
    dest_file_path : str | Path
        Ruta al libro destino.
    sheet_name : str
        Nombre de la pestana a verificar.
    month : str
        Mes objetivo.
    expected_dict : dict
        Diccionario ``{indicador: valor_esperado}`` de referencia.
    data_start_row : int
        Primera fila de datos. Default: 6.
    indicator_col : int
        Columna del indicador. Default: 1.

    Returns
    -------
    dict
        Reporte con claves: ``ok`` (bool), ``matches``, ``mismatches``, ``missing``.
    """
    dest_path = Path(dest_file_path)
    wb = openpyxl.load_workbook(dest_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    real_col = _find_real_column(ws, month)

    matches: dict = {}
    mismatches: dict = {}
    missing: list = []

    found_in_sheet: dict = {}
    for row in range(data_start_row, ws.max_row + 1):
        ind_val = ws.cell(row=row, column=indicator_col).value
        if ind_val is None:
            continue
        try:
            ind_key = int(ind_val)
        except (TypeError, ValueError):
            continue
        found_in_sheet[ind_key] = ws.cell(row=row, column=real_col).value

    wb.close()

    for ind, expected_val in expected_dict.items():
        if ind not in found_in_sheet:
            missing.append(ind)
        elif found_in_sheet[ind] == expected_val:
            matches[ind] = expected_val
        else:
            mismatches[ind] = {
                "expected": expected_val,
                "got": found_in_sheet[ind],
            }

    return {
        "ok": len(mismatches) == 0 and len(missing) == 0,
        "matches": matches,
        "mismatches": mismatches,
        "missing_indicators": missing,
        "total_checked": len(expected_dict),
    }
