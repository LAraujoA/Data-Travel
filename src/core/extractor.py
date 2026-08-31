"""
extractor.py — Fase 1 Core: Extraccion de datos mensuales desde archivos origen.

Responsabilidad:
- Dado un archivo Excel de unidad de salud origen y un mes objetivo (ej: "ENERO"),
  localizar la columna REAL correspondiente y devolver el diccionario
  {numero_indicador: valor_real}.

Estructura esperada del archivo:
  Fila 4: encabezados principales  -> col A: "No", col B: "ACTIVIDADES PRIORIZADAS",
           col C en adelante: nombre del mes (ej: "ENERO") seguido de None, None...
  Fila 5: sub-encabezados          -> "PROG", "REAL", "%"
  Fila 6+: datos                   -> col A: numero indicador (int), cols C+: valores
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

import openpyxl


def _find_month_columns(ws, target_month: str) -> tuple[int, int]:
    """
    Localiza las columnas PROG y REAL para el mes indicado.

    Escanea la fila 4 buscando ``target_month`` (case-insensitive) y luego
    busca "REAL" en la fila 5 dentro del bloque de columnas del mes.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Hoja de calculo activa del archivo origen.
    target_month : str
        Nombre del mes a buscar, ej: "ENERO", "FEBRERO".

    Returns
    -------
    tuple[int, int]
        (col_prog, col_real) — indices de columna 1-based.

    Raises
    ------
    ValueError
        Si el mes o la sub-columna REAL no se encuentran.
    """
    target_month = target_month.strip().upper()

    # --- Paso 1: encontrar la columna de inicio del mes en fila 4 ---
    month_col_start: int | None = None
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell_val = ws.cell(row=4, column=col).value
        if cell_val and str(cell_val).strip().upper() == target_month:
            month_col_start = col
            break

    if month_col_start is None:
        raise ValueError(
            f"Mes '{target_month}' no encontrado en la fila 4 del archivo."
        )

    # --- Paso 2: dentro del bloque del mes, buscar "REAL" en fila 5 ---
    # El bloque del mes ocupa columnas [month_col_start, month_col_start + N)
    # hasta que otra celda no-None aparezca en fila 4 (o fin de hoja).
    real_col: int | None = None
    prog_col: int | None = None

    for col in range(month_col_start, max_col + 1):
        # Comprobar si entramos en otro mes (celda no nula en fila 4)
        if col > month_col_start:
            next_header = ws.cell(row=4, column=col).value
            if next_header is not None:
                break  # salimos del bloque del mes actual

        sub_header = ws.cell(row=5, column=col).value
        if sub_header is not None:
            sub_header_norm = str(sub_header).strip().upper()
            if sub_header_norm == "PROG" and prog_col is None:
                prog_col = col
            elif sub_header_norm == "REAL" and real_col is None:
                real_col = col

    if real_col is None:
        raise ValueError(
            f"Sub-columna 'REAL' no encontrada para el mes '{target_month}'."
        )

    prog_col = prog_col or month_col_start  # fallback
    return prog_col, real_col


def extract_month_data(
    filepath: Union[str, Path],
    month: str,
    header_row: int = 4,
    data_start_row: int = 6,
    indicator_col: int = 1,
) -> dict:
    """
    Lee un archivo Excel origen y extrae {numero_indicador: valor_real}
    para el mes especificado.

    Parameters
    ----------
    filepath : str | Path
        Ruta al archivo .xlsx de la unidad de salud origen.
    month : str
        Nombre del mes objetivo, ej: ``"ENERO"``, ``"FEBRERO"``.
    header_row : int
        Fila (1-based) donde estan los encabezados de mes. Default: 4.
    data_start_row : int
        Primera fila de datos (1-based). Default: 6.
    indicator_col : int
        Columna (1-based) que contiene el numero de indicador. Default: 1.

    Returns
    -------
    dict
        Diccionario ``{int_indicador: valor_real}`` donde los indicadores
        son enteros y los valores son los leidos de la celda REAL.
        Filas con indicador nulo o no-numerico se omiten silenciosamente.

    Raises
    ------
    FileNotFoundError
        Si el archivo no existe.
    ValueError
        Si el mes o la columna REAL no se encuentran.

    Examples
    --------
    >>> data = extract_month_data("EL CUCO.xlsx", "ENERO")
    >>> data
    {1: 4, 2: 12, 3: 80, ...}
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    _, real_col = _find_month_columns(ws, month)

    result: dict = {}
    for row in range(data_start_row, ws.max_row + 1):
        indicator_val = ws.cell(row=row, column=indicator_col).value
        real_val = ws.cell(row=row, column=real_col).value

        # Omitir filas vacias o de totales (indicador no numerico)
        if indicator_val is None:
            continue
        try:
            indicator_key = int(indicator_val)
        except (TypeError, ValueError):
            continue

        result[indicator_key] = real_val

    wb.close()
    return result


def extract_month_data_full(
    filepath: Union[str, Path],
    month: str,
    data_start_row: int = 6,
    indicator_col: int = 1,
) -> list:
    """
    Version extendida: devuelve lista de dicts con indicador, prog y real.

    Returns
    -------
    list[dict]
        Lista de ``{"indicador": int, "prog": valor, "real": valor}``.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    prog_col, real_col = _find_month_columns(ws, month)

    rows_data = []
    for row in range(data_start_row, ws.max_row + 1):
        indicator_val = ws.cell(row=row, column=indicator_col).value
        if indicator_val is None:
            continue
        try:
            indicator_key = int(indicator_val)
        except (TypeError, ValueError):
            continue

        rows_data.append(
            {
                "indicador": indicator_key,
                "prog": ws.cell(row=row, column=prog_col).value,
                "real": ws.cell(row=row, column=real_col).value,
            }
        )

    wb.close()
    return rows_data
