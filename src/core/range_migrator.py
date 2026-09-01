
"""
range_migrator.py — Fase 4: Extraccion y distribucion de rangos libres.

Responsabilidades:
  - Leer un rango arbitrario de un archivo Excel origen (ej: "C3:C13").
  - Distribuir la matriz resultante al destino segun tres modalidades:
      a) Bloque Continuo   : pega la matriz a partir de una celda de inicio.
      b) Patron con Salto  : distribuye valores 1-D saltando N columnas o filas.
      c) Lista de Celdas   : asigna cada valor a una celda destino explicita.
  - Soportar destinos Excel local (openpyxl) y Google Sheets (gspread).
"""

from __future__ import annotations

import logging
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Union

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
#  Helpers de notacion A1
# ─────────────────────────────────────────────────────────────

_A1_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _parse_cell(cell: str) -> tuple[int, int]:
    """
    Convierte notacion A1 (ej: 'C10') a (row_1based, col_1based).

    Parameters
    ----------
    cell : str
        Referencia de celda en notacion A1, ej: 'C10', 'AA3'.

    Returns
    -------
    tuple[int, int]
        (fila, columna) ambos 1-based.

    Raises
    ------
    ValueError
        Si la cadena no tiene formato valido de celda A1.
    """
    m = _A1_RE.match(cell.strip())
    if not m:
        raise ValueError(f"Celda invalida: '{cell}'. Usa notacion A1, ej: 'C10'.")
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))
    return row, col


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """
    Convierte un rango A1 (ej: 'C3:C13') a (row1, col1, row2, col2), 1-based.

    Tambien acepta celdas sueltas (ej: 'C3') interpretandolas como 1x1.

    Parameters
    ----------
    range_str : str
        Rango en notacion A1, ej: 'C3:C13', 'A1:D10', 'B5'.

    Returns
    -------
    tuple[int, int, int, int]
        (row_min, col_min, row_max, col_max) 1-based.
    """
    parts = range_str.strip().upper().split(":")
    if len(parts) == 1:
        r, c = _parse_cell(parts[0])
        return r, c, r, c
    if len(parts) == 2:
        r1, c1 = _parse_cell(parts[0])
        r2, c2 = _parse_cell(parts[1])
        return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)
    raise ValueError(f"Rango invalido: '{range_str}'")


def _cell_a1(row: int, col: int) -> str:
    """Convierte (row, col) 1-based a notacion A1."""
    return f"{get_column_letter(col)}{row}"


def _expand_cell_tokens(cell_list_str: str) -> list[str]:
    """
    Parsea un string de destinos separados por coma y devuelve
    una lista plana de referencias de celda A1.

    Acepta cualquier combinacion de:
      - Celdas simples : "C20, G20"
      - Rangos         : "C20:C22, G20:G22"
      - Mixta          : "C20, D20:D22, H5"

    Parameters
    ----------
    cell_list_str : str
        String con celdas/rangos separados por coma.

    Returns
    -------
    list[str]
        Lista plana de celdas A1 en mayusculas, ej: ['C20','C21','C22','G20'].

    Raises
    ------
    ValueError
        Si algun token no es celda ni rango A1 valido.

    Examples
    --------
    >>> _expand_cell_tokens("C20, C21, C22")
    ['C20', 'C21', 'C22']
    >>> _expand_cell_tokens("C20:C22, G20:G22")
    ['C20', 'C21', 'C22', 'G20', 'G21', 'G22']
    """
    tokens = [t.strip().upper() for t in cell_list_str.split(",") if t.strip()]
    if not tokens:
        raise ValueError("cell_list_str esta vacia o no contiene celdas validas.")

    result: list[str] = []
    for tok in tokens:
        if ":" in tok:
            # Es un rango — expandir a todas las celdas que lo componen
            r1, c1, r2, c2 = _parse_range(tok)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    result.append(_cell_a1(r, c))
        else:
            # Es una celda simple — validar y agregar
            _parse_cell(tok)   # lanza ValueError si es invalida
            result.append(tok)
    return result


# ─────────────────────────────────────────────────────────────
#  Extraccion
# ─────────────────────────────────────────────────────────────

def extract_range(
    filepath: Union[str, Path],
    range_str: str,
    sheet_name: str | None = None,
) -> list[list]:
    """
    Lee un rango de un archivo Excel y devuelve una matriz de valores.

    Parameters
    ----------
    filepath : str | Path
        Ruta al archivo .xlsx origen.
    range_str : str
        Rango en notacion A1, ej: 'C3:C13', 'A2:D10'.
    sheet_name : str | None
        Nombre de la hoja. Si es None usa la hoja activa.

    Returns
    -------
    list[list]
        Matriz 2D ``[[fila1col1, fila1col2, ...], [fila2col1, ...], ...]``.
        Una sola columna devuelve N listas de 1 elemento cada una.

    Raises
    ------
    FileNotFoundError
        Si el archivo no existe.
    KeyError
        Si la hoja no existe en el libro.
    ValueError
        Si el rango tiene formato invalido.

    Examples
    --------
    >>> data = extract_range("reporte.xlsx", "C3:C13")
    >>> # [[4], [12], [80], ...]   <- columna C, filas 3-13
    >>> data = extract_range("reporte.xlsx", "A2:D5")
    >>> # [[v1, v2, v3, v4], [v5, v6, v7, v8], ...]
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    row1, col1, row2, col2 = _parse_range(range_str)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise KeyError(
                f"Hoja '{sheet_name}' no existe. "
                f"Disponibles: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    matrix: list[list] = []
    for r in range(row1, row2 + 1):
        row_data = []
        for c in range(col1, col2 + 1):
            row_data.append(ws.cell(row=r, column=c).value)
        matrix.append(row_data)

    wb.close()
    return matrix


def get_sheet_names(filepath: Union[str, Path]) -> list[str]:
    """Devuelve la lista de nombres de hojas de un archivo Excel."""
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ─────────────────────────────────────────────────────────────
#  Matching de hoja destino (reutiliza matcher.py)
# ─────────────────────────────────────────────────────────────

def match_dest_sheet(
    origin_label: str,
    dest_sheet_names: list[str],
    score_cutoff: float = 60.0,
) -> tuple[str | None, float]:
    """
    Encuentra la hoja destino cuyo nombre normalizado mejor coincide con el
    origen, usando la logica existente de matcher.py (rapidfuzz).

    Parameters
    ----------
    origin_label : str
        Nombre del archivo origen (ej: 'EL CUCO.xlsx') o nombre de hoja
        (ej: 'US-B Carolina SM La Ceibita').
    dest_sheet_names : list[str]
        Hojas disponibles en el libro destino.
    score_cutoff : float
        Umbral minimo de similitud (0-100). Default 60.

    Returns
    -------
    tuple[str | None, float]
        (nombre_hoja_ganadora, puntuacion) o (None, 0.0) si no hay match.

    Examples
    --------
    >>> match_dest_sheet("EL CUCO.xlsx", ["ElCuco", "SanPedro", "LaCeibita"])
    ('ElCuco', 100.0)
    >>> match_dest_sheet("US-B Carolina SM La Ceibita.xlsx", ["LaCeibita", ...])
    ('LaCeibita', 100.0)
    """
    try:
        from src.core.matcher import match_file_to_sheet
        return match_file_to_sheet(origin_label, dest_sheet_names, score_cutoff)
    except ImportError:
        log.warning(
            "matcher.py no disponible; usando comparacion exacta normalizada.")
        # Fallback: normalizacion basica sin rapidfuzz
        def _norm(s: str) -> str:
            import unicodedata, re
            nfd = unicodedata.normalize("NFD", s)
            s = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
            s = re.sub(r"[^a-z0-9]", "", s.lower())
            return s
        q = _norm(origin_label)
        for name in dest_sheet_names:
            if _norm(name) == q:
                return name, 100.0
        return None, 0.0


# ─────────────────────────────────────────────────────────────
#  Extraccion multi-rango (expresion compuesta)
# ─────────────────────────────────────────────────────────────

def extract_multi_range(
    filepath,
    range_expr: str,
    sheet_name=None,
) -> list:
    """
    Extrae valores de una expresion de rango compuesta y devuelve lista plana.

    Acepta:
      - Un rango simple         : "C3:C13"
      - Celdas individuales     : "C3, C4, C5"
      - Rangos multiples        : "C3:C5, G3:G5"
      - Mixta                   : "C3, D3:D5, H6"

    El proceso es:
      1. Dividir por comas -> lista de tokens.
      2. Por cada token: si contiene ':' leer como rango; si no, leer celda.
      3. Concatenar todos los valores en una lista plana.

    Parameters
    ----------
    filepath : str | Path
        Archivo .xlsx origen.
    range_expr : str
        Expresion de rango compuesta, ej: "C3:C5, G3:G5".
    sheet_name : str | None
        Nombre de la hoja. None = hoja activa.

    Returns
    -------
    list
        Lista plana de valores en el orden de los tokens.

    Examples
    --------
    >>> extract_multi_range("report.xlsx", "D6:D8, G6:G8")
    [4, 12, 80, 5, 25, 3]
    >>> extract_multi_range("report.xlsx", "D6, D7, D8")
    [4, 12, 80]
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    tokens = [t.strip().upper() for t in range_expr.split(",") if t.strip()]
    if not tokens:
        raise ValueError(f"Expresion de rango vacia: '{range_expr}'")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise KeyError(
                f"Hoja '{sheet_name}' no existe. "
                f"Disponibles: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    result: list = []
    for tok in tokens:
        if ":" in tok:
            r1, c1, r2, c2 = _parse_range(tok)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    result.append(ws.cell(row=r, column=c).value)
        else:
            r, c = _parse_cell(tok)
            result.append(ws.cell(row=r, column=c).value)

    wb.close()
    return result



# ─────────────────────────────────────────────────────────────
#  Distribucion — modalidad A: Bloque Continuo
# ─────────────────────────────────────────────────────────────

def write_block(
    dest_path: Union[str, Path],
    sheet_name: str,
    matrix: list[list],
    start_cell: str,
    create_backup: bool = True,
) -> dict:
    """
    Escribe la matriz en bloque continuo a partir de start_cell.

    La fila 0 de la matriz va en la fila de start_cell, la columna 0
    va en la columna de start_cell. Escala automaticamente.

    Parameters
    ----------
    dest_path : str | Path
        Ruta al archivo destino.
    sheet_name : str
        Nombre de la hoja destino.
    matrix : list[list]
        Matriz 2D de valores a escribir.
    start_cell : str
        Celda de inicio en A1, ej: 'B5'.
    create_backup : bool
        Si True crea un backup .backup.xlsx antes de guardar.

    Returns
    -------
    dict
        ``{"written": int, "cells": list[str], "backup": str | None}``

    Raises
    ------
    FileNotFoundError, KeyError, PermissionError
    """
    dest_path = Path(dest_path)
    if not dest_path.exists():
        raise FileNotFoundError(f"Destino no encontrado: {dest_path}")

    backup = None
    if create_backup:
        backup = dest_path.with_suffix(".backup.xlsx")
        shutil.copy2(dest_path, backup)

    start_row, start_col = _parse_cell(start_cell)

    wb = openpyxl.load_workbook(dest_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Hoja '{sheet_name}' no existe. Disponibles: {wb.sheetnames}")
    ws = wb[sheet_name]

    written = 0
    cells_written: list[str] = []
    detail_pairs: list[str] = []
    for r_off, row in enumerate(matrix):
        for c_off, val in enumerate(row):
            tgt_row = start_row + r_off
            tgt_col = start_col + c_off
            coord = _cell_a1(tgt_row, tgt_col)
            ws.cell(row=tgt_row, column=tgt_col).value = val
            cells_written.append(coord)
            detail_pairs.append(f"{coord}={val}")
            written += 1

    wb.save(dest_path)
    wb.close()
    detail_str = ", ".join(detail_pairs[:10])
    if written > 10:
        detail_str += f" ... (+{written - 10} mas)"
    log.info(
        "[Bloque] '%s' -> %d celdas desde %s: %s",
        sheet_name, written, start_cell, detail_str,
    )
    return {
        "written": written,
        "cells": cells_written,
        "detail": detail_pairs,
        "backup": str(backup) if backup else None,
    }


# ─────────────────────────────────────────────────────────────
#  Distribucion — modalidad B: Patron con Salto (Stride)
# ─────────────────────────────────────────────────────────────

def write_stride(
    dest_path: Union[str, Path],
    sheet_name: str,
    values: list,
    start_cell: str,
    direction: str = "Horizontal",
    stride: int = 1,
    create_backup: bool = True,
) -> dict:
    """
    Escribe valores 1-D saltando N columnas o filas desde start_cell.

    El valor values[0] va en start_cell, values[1] en start_cell + stride,
    values[2] en start_cell + 2*stride, etc.

    Parameters
    ----------
    dest_path : str | Path
        Ruta al archivo destino.
    sheet_name : str
        Nombre de la hoja destino.
    values : list
        Lista plana de valores. Si la matriz es Nx1 o 1xN, se aplana auto.
    start_cell : str
        Celda de inicio, ej: 'C20'.
    direction : str
        'Horizontal' avanza por columnas, 'Vertical' por filas.
    stride : int
        Numero de columnas/filas a saltar entre valores consecutivos.
        Ejemplo: stride=7 con Horizontal: C20 -> J20 -> Q20 ...
    create_backup : bool
        Si True crea backup .backup.xlsx.

    Returns
    -------
    dict
        ``{"written": int, "cells": list[str], "backup": str | None}``

    Raises
    ------
    ValueError
        Si direction no es 'Horizontal' ni 'Vertical', o stride < 1.
    """
    if direction not in ("Horizontal", "Vertical"):
        raise ValueError(
            f"direction debe ser 'Horizontal' o 'Vertical', no '{direction}'.")
    if stride < 1:
        raise ValueError(f"stride debe ser >= 1, no {stride}.")

    dest_path = Path(dest_path)
    if not dest_path.exists():
        raise FileNotFoundError(f"Destino no encontrado: {dest_path}")

    backup = None
    if create_backup:
        backup = dest_path.with_suffix(".backup.xlsx")
        shutil.copy2(dest_path, backup)

    # Aplanar si viene de extract_range (lista de listas)
    flat: list = []
    for item in values:
        if isinstance(item, list):
            flat.extend(item)
        else:
            flat.append(item)

    start_row, start_col = _parse_cell(start_cell)

    wb = openpyxl.load_workbook(dest_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Hoja '{sheet_name}' no existe. Disponibles: {wb.sheetnames}")
    ws = wb[sheet_name]

    written = 0
    cells_written: list[str] = []
    detail_pairs: list[str] = []
    for i, val in enumerate(flat):
        if direction == "Horizontal":
            tgt_row = start_row
            tgt_col = start_col + i * stride
        else:  # Vertical
            tgt_row = start_row + i * stride
            tgt_col = start_col
        coord = _cell_a1(tgt_row, tgt_col)
        ws.cell(row=tgt_row, column=tgt_col).value = val
        cells_written.append(coord)
        detail_pairs.append(f"{coord}={val}")
        written += 1

    wb.save(dest_path)
    wb.close()
    detail_str = ", ".join(detail_pairs[:10])
    if written > 10:
        detail_str += f" ... (+{written - 10} mas)"
    log.info(
        "[Stride-%s paso=%d] '%s' -> %d celdas: %s",
        direction, stride, sheet_name, written, detail_str,
    )
    return {
        "written": written,
        "cells": cells_written,
        "detail": detail_pairs,
        "backup": str(backup) if backup else None,
    }


# ─────────────────────────────────────────────────────────────
#  Distribucion — modalidad C: Lista Explicita de Celdas
# ─────────────────────────────────────────────────────────────

def write_cell_list(
    dest_path: Union[str, Path],
    sheet_name: str,
    values: list,
    cell_list_str: str,
    create_backup: bool = True,
) -> dict:
    """
    Asigna cada valor de `values` a la celda correspondiente en cell_list_str.

    Si hay mas celdas que valores, las celdas extra quedan sin tocar.
    Si hay mas valores que celdas, los valores extra se omiten con advertencia.

    Parameters
    ----------
    dest_path : str | Path
        Ruta al archivo destino.
    sheet_name : str
        Nombre de la hoja destino.
    values : list
        Lista plana de valores (o lista de listas, se aplana automaticamente).
    cell_list_str : str
        Celdas separadas por coma, ej: 'C20, G20, K20, Q20'.
    create_backup : bool
        Si True crea backup .backup.xlsx.

    Returns
    -------
    dict
        ``{"written": int, "cells": list[str], "backup": str | None}``

    Examples
    --------
    >>> write_cell_list(
    ...     "POA.xlsx", "LaCeibita",
    ...     [4, 12, 80, 5],
    ...     "C20, G20, K20, Q20",
    ... )
    {"written": 4, "cells": ["C20", "G20", "K20", "Q20"], "backup": ...}
    """
    dest_path = Path(dest_path)
    if not dest_path.exists():
        raise FileNotFoundError(f"Destino no encontrado: {dest_path}")

    # Bug 1 fix: usar _expand_cell_tokens para soportar celdas simples
    # Y rangos mezclados: "C20, C21" o "C20:C22, G20:G22"
    target_cells = _expand_cell_tokens(cell_list_str)

    # Aplanar valores (puede venir como list[list] de extract_range)
    flat: list = []
    for item in values:
        if isinstance(item, list):
            flat.extend(item)
        else:
            flat.append(item)

    if len(flat) > len(target_cells):
        log.warning(
            "Hay %d valores pero solo %d celdas destino. "
            "Los %d valores extra se omiten.",
            len(flat), len(target_cells), len(flat) - len(target_cells),
        )

    backup = None
    if create_backup:
        backup = dest_path.with_suffix(".backup.xlsx")
        shutil.copy2(dest_path, backup)

    # Bug 2 fix: abrir en modo r/w (sin data_only), acceder a la hoja
    # correcta, escribir celda a celda, save() expliciton y cerrar.
    wb = openpyxl.load_workbook(dest_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Hoja '{sheet_name}' no existe. Disponibles: {wb.sheetnames}")
    ws = wb[sheet_name]

    written = 0
    cells_written: list[str] = []
    detail_pairs: list[str] = []
    for cell_ref, val in zip(target_cells, flat):
        r, c = _parse_cell(cell_ref)
        ws.cell(row=r, column=c).value = val
        coord = cell_ref.upper()
        cells_written.append(coord)
        detail_pairs.append(f"{coord}={val}")
        written += 1

    wb.save(dest_path)
    wb.close()

    detail_str = ", ".join(detail_pairs[:10])
    if written > 10:
        detail_str += f" ... (+{written - 10} mas)"
    log.info(
        "[Lista] '%s' -> %d celdas: %s",
        sheet_name, written, detail_str,
    )
    return {
        "written": written,
        "cells": cells_written,
        "detail": detail_pairs,
        "backup": str(backup) if backup else None,
    }


# ─────────────────────────────────────────────────────────────
#  API unificada
# ─────────────────────────────────────────────────────────────

MODES = ("Bloque Continuo", "Salto de Columnas/Filas", "Lista de Celdas")


def migrate_range(
    src_file: Union[str, Path],
    src_range: str,
    dest_file: Union[str, Path],
    dest_sheet: str,
    mode: str,
    src_sheet: str | None = None,
    start_cell: str = "A1",
    direction: str = "Horizontal",
    stride: int = 1,
    cell_list: str = "",
    create_backup: bool = True,
) -> dict:
    """
    Funcion unificada: extrae un rango y lo distribuye segun la modalidad.

    Parameters
    ----------
    src_file : str | Path
        Archivo Excel origen.
    src_range : str
        Rango a leer, ej: 'C3:C13'.
    dest_file : str | Path
        Archivo Excel destino.
    dest_sheet : str
        Hoja destino.
    mode : str
        Una de: 'Bloque Continuo', 'Salto de Columnas/Filas', 'Lista de Celdas'.
    src_sheet : str | None
        Hoja origen (None = activa).
    start_cell : str
        Celda de inicio para modos Bloque y Salto. Default 'A1'.
    direction : str
        'Horizontal' o 'Vertical'. Solo para modo Salto.
    stride : int
        Paso numerico. Solo para modo Salto.
    cell_list : str
        Celdas separadas por coma. Solo para modo Lista.
    create_backup : bool
        Crear backup antes de escribir.

    Returns
    -------
    dict
        Resultado de la funcion de escritura:
        ``{"written": int, "cells": list[str], "backup": str | None}``.
        Incluye ademas ``"matrix": list[list]`` con los datos extraidos.

    Raises
    ------
    ValueError
        Si mode no es valido.
    """
    if mode not in MODES:
        raise ValueError(
            f"Modalidad '{mode}' no reconocida. Opciones: {MODES}")

    matrix = extract_range(src_file, src_range, src_sheet)
    log.info(
        "Extraido rango '%s' de '%s' -> %d filas x %d cols",
        src_range, Path(src_file).name,
        len(matrix), len(matrix[0]) if matrix else 0,
    )

    if mode == "Bloque Continuo":
        result = write_block(
            dest_file, dest_sheet, matrix,
            start_cell=start_cell,
            create_backup=create_backup,
        )
    elif mode == "Salto de Columnas/Filas":
        result = write_stride(
            dest_file, dest_sheet, matrix,
            start_cell=start_cell,
            direction=direction,
            stride=stride,
            create_backup=create_backup,
        )
    else:  # Lista de Celdas
        result = write_cell_list(
            dest_file, dest_sheet, matrix,
            cell_list_str=cell_list,
            create_backup=create_backup,
        )

    result["matrix"] = matrix
    return result
